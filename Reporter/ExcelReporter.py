import json
import os
from dataclasses import dataclass
from typing import Dict, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

import sys
import os

# Add parent directory to path to allow imports from sibling directories
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from utils.logger import ProgressLogger

@dataclass
class ExcelStyles:
    """Styles for Excel formatting"""
    header_font: Font = Font(bold=True, color="FFFFFF") 
    header_alignment: Alignment = Alignment(horizontal="center", vertical="center")
    header_fill: PatternFill = PatternFill(start_color="16365C", end_color="16365C", fill_type="solid")
    header_border: Border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )
    data_fill_even: PatternFill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    data_fill_odd: PatternFill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    data_alignment: Alignment = Alignment(horizontal="center", vertical="center")
    data_font: Font = Font(color="000000")

class ExcelReporter:
    def __init__(self, excelOutPutName: str, input_path: str, progress_logger: ProgressLogger):
        self.input_path = input_path
        self.Excel_output_path = os.path.join(input_path, excelOutPutName)
        self.DefectsCodeDict = {}
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Condition Details"
        self.styles = ExcelStyles()
        self.progress_logger = progress_logger
        self.headers = [
            "زمان بر روی ویدئو", "مسیر تصویر", "فاصله از نقطه شروع", "عیوب پیوسته", "کد عیوب",
            "نام فارسی عیوب", "عیب در محل اتصال", "جنس", "شدت عیب", "ابعاد یک", "ابعاد دو",
            "درصد", "محل قرارگیری از ساعت", "محل قرارگیری تا ساعت", "امتیاز بهره برداری",
            "امتیاز سازه ای", "ملاحضات"
        ]
        self.data_mapping = {
            "زمان بر روی ویدئو": "timestamp_seconds",
            "مسیر تصویر": "frame_path",
            "فاصله از نقطه شروع": "text_info",
            "کد عیوب": ["Detection", "class"],
            "شدت عیب": ["Detection", "confidence"]
        }

    def load_json_data(self, path) -> Dict:
        """Load and return JSON data from file"""
        with open(path, 'r', encoding='utf-8') as file:
            return json.load(file)

    def apply_header_styling(self) -> None:
        """Apply styling to header row"""
        # Set row height to 30
        self.sheet.row_dimensions[1].height = 30
        
        for col_num, header in enumerate(self.headers, 1):
            cell = self.sheet.cell(row=1, column=col_num, value=header)
            cell.font = self.styles.header_font
            cell.alignment = self.styles.header_alignment
            cell.fill = self.styles.header_fill 
            cell.border = self.styles.header_border

    def format_timestamp(self, seconds: float) -> str:
        """Convert seconds to HH:MM:SS format"""
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        seconds = int(seconds % 60)
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"

    def write_data_rows(self, json_data: Dict) -> None:
        """Write and style data rows"""
        total_rows = len(json_data)
        row_num = 2

        try:
            for _, (image_key, entry) in enumerate(json_data.items(), 0):
                # Update progress for excel reporting stage
                progress = ((row_num - 1) / total_rows) * 100
                if row_num % 10 == 0:
                    self.progress_logger.update_stage_progress(
                        "excel reporting",
                        progress,
                        {"status": f"Writing row {row_num-1} of {total_rows}"}
                    )
                
                fill = self.styles.data_fill_even if row_num % 2 == 0 else self.styles.data_fill_odd
                
                # Check if Detection is a list and process each detection as a separate row
                detections = entry.get("Detection", [])
                if not isinstance(detections, list):
                    detections = [detections]
                
                # If there are multiple detections, we'll add rows for each one
                for detection_idx, detection in enumerate(detections, 0):
                    current_row = row_num + detection_idx
                    
                    # Create a new row if needed (after the first detection)
                    if detection_idx > 0:
                        # Apply the alternating fill pattern
                        fill = self.styles.data_fill_even if current_row % 2 == 0 else self.styles.data_fill_odd
                    
                    for col_num, header in enumerate(self.headers, 1):
                        cell = self.sheet.cell(row=current_row, column=col_num)
                        key = self.data_mapping.get(header)
                        
                        if key:
                            if header == "زمان بر روی ویدئو":
                                value = self.format_timestamp(float(detection[key]))
                            elif header == "فاصله از نقطه شروع":
                                value = entry.get(key, [])
                            elif header == "کد عیوب":
                                raw_value = detection[key[1]]
                                value = self.DefectsCodeDict.get(raw_value, raw_value)
                            elif header == "شدت عیب":
                                value = detection[key[1]]
                            elif header == "نام فارسی عیوب":
                                raw_value = detection[key[1]]
                                value = self.DefectsCodeDict["BasicName_fa"].get(raw_value, raw_value)
                            else:
                                # For other fields, check if they're in detection or entry
                                if isinstance(key, list):
                                    value = detection.get(key[1], entry.get(key, ""))
                                else:
                                    value = detection.get(key, entry.get(key, ""))
                        else:
                            value = ""
                        
                        if header == "نام فارسی عیوب":
                            key = self.data_mapping.get("کد عیوب")
                            raw_value = detection[key[1]]
                            value = self.DefectsCodeDict["BasicName_fa"].get(raw_value, raw_value)
                        
                        cell.value = value
                        cell.font = self.styles.data_font
                        cell.alignment = self.styles.data_alignment
                        cell.border = self.styles.header_border
                        cell.fill = fill
                
                # Adjust row_num to account for the added rows
                row_num += len(detections)
        except Exception as e:
            # Mark excel reporting stage as complete
            self.progress_logger.complete_stage(
                "excel reporting",
                {"status": "Finished writing all data rows"}
            )
            return

        self.progress_logger.update_stage_progress(
                    "excel reporting",
                    100,
                    {"status": f"Finished writing all data rows"}
                )   
        # Mark excel reporting stage as complete
        self.progress_logger.complete_stage(
            "excel reporting",
            {"status": "Finished writing all data rows"}
        )

    def adjust_column_widths(self) -> None:
        """Adjust column widths based on content"""
        for col in self.sheet.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            self.sheet.column_dimensions[col[0].column_letter].width = max_length + 2

    def createCodeMapData(self, defects_data: Dict) -> None:
        # Extract BasicName-to-Key mapping
        basic_to_key_mapping = {}
        basic_defect_codes = {}
        basic_defect_codes["BasicName_fa"] = {
            "Root": "ریشه درخت",
            "Infiltration": "نفوذ آب", 
            "Leak": "نشت آب",
            "Broken": "شکستگی",
            "Deposits": "رسوب",
            "Infiltration of soil": "نفوذ خاک",
            "Obstacle": "مانع",
            "Water level": "تراز آب",
            "Crack": "ترک",
            "Fracture": "شکاف",
            "Hole": "سوراخ",
            "Deformed": "تغییر شکل",
            "Open connection": "اتصال باز",
            "Surface Damage": "خرابی سطج",
            "Joint Displaced": "جابجایی محل اتصال"
        }
        # Add basic defect codes to mapping
        basic_to_key_mapping.update(basic_defect_codes)
        
        # Add mappings from defects data
        for defect_category in defects_data.values():
            if isinstance(defect_category, dict) and "BasicDefects" in defect_category:
                for key, details in defect_category.items():
                    if isinstance(details, dict) and "BasicName" in details:
                        basic_to_key_mapping[details["BasicName"]] = key


        self.DefectsCodeDict = basic_to_key_mapping
        

    def generate_report(self) -> None:
        """Generate the Excel report"""
        #detectionDataPath = r"C:\Users\sobha\Desktop\detectron2\Code\Auto_Sewer_Document\output\Closed circuit television (CCTV) sewer inspection_detections.json"
        #detectionDataPath = r"C:\Users\sobha\Desktop\detectron2\Code\Auto_Sewer_Document\output\olympic-St25zdo494Surveyupstream_detections.json"

        #defectCodePath = "DefectsCode.json"
        # Get parent folder of detection data path
        base_name = os.path.basename(self.input_path).split("_")[0]
        detection_data_dir = os.path.join(self.input_path, f"frames_detections.json")

        #defectCodePath = "output/DefectsCode.json"
        # Get the directory where the current script is located
        # Get the parent directory (project root) by going up one level from the current file
        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

        # Build the path to 'model/v2'
        defectCodePath = os.path.join(BASE_DIR, 'output', 'DefectsCode.json')

        detectionData = self.load_json_data(detection_data_dir)
        defectCodeData = self.load_json_data(defectCodePath)

        self.apply_header_styling()
        self.createCodeMapData(defectCodeData)
        self.write_data_rows(detectionData)
        self.adjust_column_widths()
        try:
            self.workbook.save(self.Excel_output_path)
        except PermissionError:
            self.progress_logger.log_message("Excel file is open. Attempting to close it...")
            try:
                # Close the workbook if it's open
                if hasattr(self, '_archive'):
                    self.workbook.close()
                # Try saving again
                self.workbook.save(self.Excel_output_path)
                self.progress_logger.log_message("Successfully saved Excel file after closing.")
            except Exception as e:
                self.progress_logger.log_message(f"Error saving Excel file: {str(e)}")
                raise
        print(f"Excel file with modern styling created: {self.Excel_output_path}")

if __name__ == "__main__":
    reporter = ExcelReporter(
        "modern_styled_condition_details_filled.xlsx",
        r"C:\Users\sobha\Desktop\detectron2\Data\E.Hormozi\olympicSt25zdo4931Surveyupstream_output",
        ProgressLogger()
    )
    reporter.generate_report()
